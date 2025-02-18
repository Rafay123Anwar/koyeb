
from django.shortcuts import render, redirect
from django.core.mail import EmailMessage
from django.http import HttpResponse
from .models import EmailNameData, Certificate, Coordinate
from .forms import UploadEmailFileForm, UploadCertificateForm
import base64
from django.conf import settings
from django.db import transaction
from concurrent.futures import ThreadPoolExecutor
import fitz



def get_session_id(request):
    """Ensure a unique session ID exists for the user."""
    if not request.session.session_key:
        request.session.create()
    return request.session.session_key


def upload_email_file(request):
    session_id = get_session_id(request)

    if request.method == 'POST':
        form = UploadEmailFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            decoded_file = None

            try:
                # Handle CSV and Excel files (without pandas for lighter approach)
                if file.name.endswith(('.xls', '.xlsx', '.xlsm')):
                    from openpyxl import load_workbook
                    wb = load_workbook(file)
                    sheet = wb.active
                    decoded_file = [
                        (row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2)
                        if row[0].value and row[1].value  # Ensure both name and email are present
                    ]
                else:
                    decoded_file = [
                        (line.decode('utf-8').split(',')[0], line.decode('utf-8').split(',')[1])
                        for line in file.readlines() if line.decode('utf-8').split(',')[0] and line.decode('utf-8').split(',')[1]
                    ]

                # Pre-fetch existing emails for faster filtering
                existing_emails = set(
                    EmailNameData.objects.filter(session_id=session_id).values_list('email', flat=True)
                )

                # Bulk insert email and name data
                rows = [
                    EmailNameData(name=row[0], email=row[1], session_id=session_id)
                    for row in decoded_file if row[1] not in existing_emails and row[0]
                ]
                if rows:
                    with transaction.atomic():
                        EmailNameData.objects.bulk_create(rows)
            except Exception as e:
                print(f"Error processing file: {e}")
                return HttpResponse("An error occurred while processing the file.", status=500)

            return redirect('upload_certificate')
    else:
        form = UploadEmailFileForm()

    return render(request, 'upload_email_file.html', {'form': form})



def upload_certificate(request):
    session_id = get_session_id(request)

    if request.method == 'POST':
        form = UploadCertificateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file = request.FILES['file']
                with transaction.atomic():
                    Certificate.objects.create(file=file.read(), session_id=session_id)
                return redirect('set_coordinates')
            except Exception as e:
                print(f"Error saving certificate: {e}")
                return HttpResponse("An error occurred while uploading the certificate.", status=500)
    else:
        form = UploadCertificateForm()

    return render(request, 'upload_certificate.html', {'form': form})

def set_coordinates(request):
    session_id = get_session_id(request)
    certificate_image_data = None
    certificate_width, certificate_height = 1000, 1000  # Default dimensions

    try:
        # Fetch the latest certificate for the session
        certificate = Certificate.objects.filter(session_id=session_id).latest('uploaded_at')

        # Convert PDF to an image for display
        pdf_data = bytes(certificate.file)
        pdf_document = fitz.open(stream=pdf_data, filetype="pdf")
        page = pdf_document[0]
        pix = page.get_pixmap()
        image_data = pix.tobytes("png")
        certificate_image_data = base64.b64encode(image_data).decode('utf-8')

        certificate_width = pix.width
        certificate_height = pix.height
    except Certificate.DoesNotExist:
        return HttpResponse("Certificate not found for this session.", status=404)
    except Exception as e:
        print(f"Error processing certificate: {e}")
        return HttpResponse("An error occurred while processing the certificate.", status=500)

    if request.method == 'POST':
        try:
            x = float(request.POST.get('x'))
            y = float(request.POST.get('y'))
            font_size = int(request.POST.get('fontSize'))
            font_color = request.POST.get('fontColor')

            # Ensure that a valid certificate object is passed to the Coordinate
            with transaction.atomic():
                Coordinate.objects.create(
                    x=x,
                    y=y,
                    font_size=font_size,
                    font_color=font_color,
                    certificate=certificate,  # Assign the certificate object here
                    session_id=session_id
                )
            return redirect('send_emails')
        except Exception as e:
            print(f"Error saving coordinates: {e}")
            return HttpResponse("An error occurred while saving the coordinates.", status=500)

    return render(request, 'set_coordinates.html', {
        'certificate_image_data': certificate_image_data,
        'certificate_width': certificate_width,
        'certificate_height': certificate_height
    })



def send_email_batch(emails_data, certificate_binary, coordinate):
    for recipient in emails_data:
        font_color_rgb = tuple(int(coordinate.font_color[i:i + 2], 16) for i in (1, 3, 5))  # Optimized color conversion
        modified_pdf_data = add_name_to_certificate(
            certificate_binary=certificate_binary,
            name=recipient.name,
            x=coordinate.x,
            y=coordinate.y,
            font_size=coordinate.font_size,
            font_color=font_color_rgb
        )

        email = EmailMessage(
            "🎉 Your Personalized Certificate is Ready! 🎓",
            f"Hi {recipient.name},\n\nYour certificate is ready!",
            'noreply@example.com',
            [recipient.email]
        )
        email.attach('certificate.pdf', modified_pdf_data, 'application/pdf')
        email.send()


def send_emails(request):
    session_id = get_session_id(request)

    try:
        certificate = Certificate.objects.filter(session_id=session_id).latest('uploaded_at')
        coordinate = Coordinate.objects.filter(session_id=session_id).first()

        if not coordinate:
            return HttpResponse("Coordinates not found for this session.", status=404)

        recipients = EmailNameData.objects.filter(session_id=session_id)

        # Using ThreadPoolExecutor for parallel email sending
        with ThreadPoolExecutor(max_workers=10) as executor:
            executor.submit(send_email_batch, recipients, certificate.file, coordinate)

        return redirect('success')
    except Exception as e:
        print(f"Error sending emails: {e}")
        return HttpResponse("An error occurred while sending emails.", status=500)

def add_name_to_certificate(certificate_binary, name, x, y, font_size, font_color, font_name="MonteCarlo"):
    """Add a name to a certificate PDF at specified coordinates."""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from PyPDF2 import PdfReader, PdfWriter
        import io
        import os
        from django.conf import settings

        # Register font if not already registered
        font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "MonteCarlo-Regular.ttf")
        pdfmetrics.registerFont(TTFont(font_name, font_path if os.path.exists(font_path) else "Helvetica"))

        reader = PdfReader(io.BytesIO(certificate_binary))
        writer = PdfWriter()
        first_page = reader.pages[0]
        page_width = float(first_page.mediabox.width)
        page_height = float(first_page.mediabox.height)
        y_inverted = page_height - y

        packet = io.BytesIO()
        can = canvas.Canvas(packet, pagesize=(page_width, page_height))
        can.setFont(font_name, font_size)
        can.setFillColorRGB(*[c / 255 for c in font_color])
        can.drawString(x, y_inverted, name)
        can.save()
        packet.seek(0)

        overlay_pdf = PdfReader(packet)
        for page in reader.pages:
            page.merge_page(overlay_pdf.pages[0])
            writer.add_page(page)

        output = io.BytesIO()
        writer.write(output)
        return output.getvalue()
    except Exception as e:
        print(f"Error adding name to certificate: {e}")
        raise


def success_view(request):
    session_id = get_session_id(request)
    try:
        EmailNameData.objects.filter(session_id=session_id).delete()
        Coordinate.objects.filter(session_id=session_id).delete()
        Certificate.objects.filter(session_id=session_id).delete()
    except Exception as e:
        print(f"Error during cleanup: {e}")
    return render(request, 'success.html')
